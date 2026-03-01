"""
FMEA Generator Module
Orchestrates the complete FMEA generation pipeline
"""

import copy
import pandas as pd
import numpy as np
from typing import Dict, List, Optional, Union, Any
from pathlib import Path
import logging
from datetime import datetime

from preprocessing import DataPreprocessor
from llm_extractor import LLMExtractor
from risk_scoring import RiskScoringEngine
from multi_model_comparison import MultiModelComparator

logger = logging.getLogger(__name__)


class FMEAGenerator:
    """
    Complete FMEA generation system
    Orchestrates preprocessing, extraction, and risk scoring
    """
    
    def __init__(self, config: Dict):
        """
        Initialize FMEA Generator with all components
        
        Args:
            config: Configuration dictionary
        """
        self.config = config
        
        logger.info("Initializing FMEA Generator components...")
        
        # Initialize modules
        self.preprocessor = DataPreprocessor(config)
        self.extractor = LLMExtractor(config)
        self.scorer = RiskScoringEngine(config)
        
        logger.info("FMEA Generator initialized successfully")
    
    def generate_multi_model_comparison(self, 
                                       text_input: Union[str, List[str]],
                                       model_names: List[str],
                                       is_file: bool = False) -> Dict[str, Any]:
        """
        Generate FMEA from multiple models and compare results
        
        Args:
            text_input: File path or list of text strings
            model_names: List of model names to use for comparison
            is_file: Whether text_input is a file path
            
        Returns:
            Dictionary containing:
            {
                'individual_results': Dict[model_name -> FMEA DataFrame],
                'comparison_results': Comprehensive comparison data from MultiModelComparator
            }
        """
        logger.info(f"Generating FMEA from {len(model_names)} models for comparison...")
        
        if len(model_names) < 2:
            raise ValueError("Need at least 2 models for comparison")
        
        # Step 1: Preprocess text (shared across all models)
        if is_file:
            preprocessed_df = self.preprocessor.load_unstructured_data(file_path=text_input)
        else:
            preprocessed_df = self.preprocessor.load_unstructured_data(text_data=text_input)
        
        texts = preprocessed_df['text_cleaned'].tolist()
        
        # Step 2: Generate FMEA for each model
        individual_results = {}
        
        for model_name in model_names:
            logger.info(f"Generating FMEA for model: {model_name}")
            
            # Thread-safe: deep copy config for each model
            model_config = copy.deepcopy(self.config)
            model_config['model']['name'] = model_name
            
            # Create new extractor with isolated config
            temp_extractor = LLMExtractor(model_config)
            
            # Extract failure information using this model
            extracted_info = temp_extractor.batch_extract(texts)
            extracted_df = pd.DataFrame(extracted_info)
            
            # Add original text for reference
            extracted_df['original_text'] = preprocessed_df['text'].values
            extracted_df['sentiment'] = preprocessed_df['sentiment'].values
            
            # Calculate risk scores
            fmea_df = self.scorer.batch_score(extracted_df)
            
            # Generate recommendations
            fmea_df = self._generate_recommendations(fmea_df)
            
            # Format output
            fmea_df = self._format_output(fmea_df)
            
            individual_results[model_name] = fmea_df
            
            logger.info(f"Generated FMEA for {model_name} with {len(fmea_df)} entries")
        
        # Step 3: Compare results from all models
        comparator = MultiModelComparator(self.config)
        comparison_results = comparator.compare_models(individual_results)
        
        logger.info("Multi-model comparison completed successfully")
        
        return {
            'individual_results': individual_results,
            'comparison_results': comparison_results
        }
    
    def generate_multi_model_from_structured(self,
                                            file_path: str,
                                            model_names: List[str]) -> Dict[str, Any]:
        """
        Generate FMEA from structured data using multiple models
        
        Args:
            file_path: Path to CSV or Excel file
            model_names: List of model names to use for comparison
            
        Returns:
            Dictionary containing individual results and comparison data
        """
        logger.info(f"Generating FMEA from structured file using {len(model_names)} models...")
        
        if len(model_names) < 2:
            raise ValueError("Need at least 2 models for comparison")
        
        # Load and validate structured data (shared across all models)
        structured_df = self.preprocessor.load_structured_data(file_path)
        
        # Check if risk scores already exist
        has_scores = all(col in structured_df.columns 
                        for col in ['severity', 'occurrence', 'detection'])
        
        individual_results = {}
        
        for model_name in model_names:
            logger.info(f"Processing structured data for model: {model_name}")
            
            # Thread-safe: deep copy config for each model
            model_config = copy.deepcopy(self.config)
            model_config['model']['name'] = model_name
            
            fmea_df = structured_df.copy()
            
            if not has_scores:
                # For comparison, we use the same scorer for all models
                # The difference would come from different extracted failure modes
                temp_scorer = RiskScoringEngine(model_config)
                fmea_df = temp_scorer.batch_score(fmea_df)
            else:
                # Recalculate RPN with updated config settings
                fmea_df['rpn'] = fmea_df.apply(
                    lambda row: self.scorer.calculate_rpn(
                        row['severity'], row['occurrence'], row['detection']
                    ), axis=1
                )
                fmea_df['action_priority'] = fmea_df.apply(
                    lambda row: self.scorer.calculate_action_priority(
                        row['severity'], row['occurrence'], row['detection']
                    ), axis=1
                )
            
            # Generate recommendations
            fmea_df = self._generate_recommendations(fmea_df)
            
            # Format output
            fmea_df = self._format_output(fmea_df)
            
            individual_results[model_name] = fmea_df
            
            logger.info(f"Processed structured data for {model_name} with {len(fmea_df)} entries")
        
        # Compare results from all models
        comparator = MultiModelComparator(self.config)
        comparison_results = comparator.compare_models(individual_results)
        
        logger.info("Multi-model comparison from structured data completed successfully")
        
        return {
            'individual_results': individual_results,
            'comparison_results': comparison_results
        }

    
    def generate_from_text(self, text_input: Union[str, List[str]], 
                          is_file: bool = False) -> pd.DataFrame:
        """
        Generate FMEA from unstructured text input
        
        Args:
            text_input: File path or list of text strings
            is_file: Whether text_input is a file path
            
        Returns:
            Complete FMEA DataFrame
        """
        logger.info("Generating FMEA from unstructured text...")
        
        # Step 1: Preprocess text
        if is_file:
            preprocessed_df = self.preprocessor.load_unstructured_data(file_path=text_input)
        else:
            preprocessed_df = self.preprocessor.load_unstructured_data(text_data=text_input)
        
        # Step 2: Extract failure information using LLM
        texts = preprocessed_df['text_cleaned'].tolist()
        extracted_info = self.extractor.batch_extract(texts)
        
        # Convert to DataFrame
        extracted_df = pd.DataFrame(extracted_info)
        
        # Add original text for reference
        extracted_df['original_text'] = preprocessed_df['text'].values
        extracted_df['sentiment'] = preprocessed_df['sentiment'].values
        
        # Step 3: Calculate risk scores
        fmea_df = self.scorer.batch_score(extracted_df)
        
        # Step 4: Generate recommended actions
        fmea_df = self._generate_recommendations(fmea_df)
        
        # Step 5: Format final output
        fmea_df = self._format_output(fmea_df)
        
        logger.info(f"Generated FMEA with {len(fmea_df)} entries")
        
        return fmea_df
    
    def generate_from_structured(self, file_path: str) -> pd.DataFrame:
        """
        Generate FMEA from structured input (CSV/Excel)
        
        Args:
            file_path: Path to CSV or Excel file
            
        Returns:
            Complete FMEA DataFrame
        """
        logger.info(f"Generating FMEA from structured file: {file_path}")
        
        # Step 1: Load and validate structured data
        result = self.preprocessor.load_structured_data(file_path)
        
        # Handle both tuple (new) and DataFrame (old) return types for backward compatibility
        if isinstance(result, tuple):
            structured_df, validation_result = result
            logger.info(f"Validation result: {validation_result.valid_records}/{validation_result.total_records} records valid")
        else:
            structured_df = result
        
        # Step 2: Check if risk scores already exist
        has_scores = all(col in structured_df.columns 
                        for col in ['severity', 'occurrence', 'detection'])
        
        if not has_scores:
            # Calculate risk scores
            logger.info("Calculating risk scores for structured data...")
            fmea_df = self.scorer.batch_score(structured_df)
        else:
            # Use existing scores, recalculate RPN
            logger.info("Using existing risk scores from file")
            fmea_df = structured_df.copy()
            fmea_df['rpn'] = fmea_df.apply(
                lambda row: self.scorer.calculate_rpn(
                    row['severity'], row['occurrence'], row['detection']
                ), axis=1
            )
            fmea_df['action_priority'] = fmea_df.apply(
                lambda row: self.scorer.calculate_action_priority(
                    row['severity'], row['occurrence'], row['detection']
                ), axis=1
            )
        
        # Step 3: Generate recommended actions
        fmea_df = self._generate_recommendations(fmea_df)
        
        # Step 4: Format output
        fmea_df = self._format_output(fmea_df)
        
        logger.info(f"Generated FMEA with {len(fmea_df)} entries")
        
        return fmea_df
    
    def generate_hybrid(self, structured_file: Optional[str] = None,
                       text_input: Optional[Union[str, List[str]]] = None) -> pd.DataFrame:
        """
        Generate FMEA from both structured and unstructured inputs
        
        Args:
            structured_file: Path to structured data file
            text_input: Unstructured text data
            
        Returns:
            Combined FMEA DataFrame
        """
        logger.info("Generating hybrid FMEA from multiple sources...")
        
        dataframes = []
        
        # Process structured data
        if structured_file:
            structured_fmea = self.generate_from_structured(structured_file)
            structured_fmea['source'] = 'Structured Data'
            dataframes.append(structured_fmea)
        
        # Process unstructured data
        if text_input:
            is_file = isinstance(text_input, str) and Path(text_input).exists()
            text_fmea = self.generate_from_text(text_input, is_file=is_file)
            text_fmea['source'] = 'Unstructured Text'
            dataframes.append(text_fmea)
        
        if not dataframes:
            raise ValueError("No input data provided")
        
        # Combine all sources
        combined_fmea = pd.concat(dataframes, ignore_index=True)
        
        # Remove duplicates based on similarity
        combined_fmea = self._deduplicate_failures(combined_fmea)
        
        # Re-sort by RPN
        combined_fmea = combined_fmea.sort_values('Rpn', ascending=False).reset_index(drop=True)
        
        logger.info(f"Generated combined FMEA with {len(combined_fmea)} entries")
        
        return combined_fmea
    
    def _generate_recommendations(self, fmea_df: pd.DataFrame) -> pd.DataFrame:
        """
        Generate recommended actions based on risk scores
        
        Args:
            fmea_df: FMEA DataFrame with risk scores
            
        Returns:
            DataFrame with added recommendations
        """
        def get_recommendation(row):
            priority = row.get('action_priority', 'Medium')
            severity = row.get('severity', 5)
            occurrence = row.get('occurrence', 5)
            detection = row.get('detection', 5)
            
            recommendations = []
            
            # Severity-based recommendations
            if severity >= 8:
                recommendations.append("Immediate design review required")
                recommendations.append("Implement redundant safety systems")
            elif severity >= 6:
                recommendations.append("Enhance safety controls")
            
            # Occurrence-based recommendations
            if occurrence >= 8:
                recommendations.append("Root cause analysis needed")
                recommendations.append("Process improvement required")
            elif occurrence >= 6:
                recommendations.append("Implement preventive maintenance")
            
            # Detection-based recommendations
            if detection >= 8:
                recommendations.append("Improve detection methods")
                recommendations.append("Add monitoring systems")
            elif detection >= 6:
                recommendations.append("Enhance inspection procedures")
            
            if priority == 'Critical':
                recommendations.insert(0, "URGENT: Immediate action required")
            
            return " | ".join(recommendations) if recommendations else "Continue monitoring"
        
        fmea_df['recommended_action'] = fmea_df.apply(get_recommendation, axis=1)
        
        return fmea_df
    
    def _format_output(self, fmea_df: pd.DataFrame) -> pd.DataFrame:
        """
        Format FMEA output with proper column order and naming
        
        Args:
            fmea_df: FMEA DataFrame
            
        Returns:
            Formatted DataFrame
        """
        # Define standard FMEA column order
        standard_columns = [
            'failure_mode',
            'effect',
            'cause',
            'component',
            'process',
            'existing_controls',
            'severity',
            'occurrence',
            'detection',
            'rpn',
            'action_priority',
            'recommended_action'
        ]
        
        # Add optional columns if they exist
        optional_columns = ['source', 'original_text', 'sentiment']
        
        # Select available columns
        output_columns = [col for col in standard_columns if col in fmea_df.columns]
        output_columns += [col for col in optional_columns if col in fmea_df.columns]
        
        # Ensure process column exists
        if 'process' not in fmea_df.columns:
            fmea_df['process'] = fmea_df.get('component', 'General Process')
        
        result_df = fmea_df[output_columns].copy()
        
        # Rename columns to proper case
        result_df.columns = [col.replace('_', ' ').title() for col in result_df.columns]
        
        # Sort by RPN (descending)
        if 'Rpn' in result_df.columns:
            result_df = result_df.sort_values('Rpn', ascending=False).reset_index(drop=True)
        
        return result_df
    
    def _deduplicate_failures(self, fmea_df: pd.DataFrame) -> pd.DataFrame:
        """
        Remove duplicate or very similar failure modes
        
        Args:
            fmea_df: FMEA DataFrame
            
        Returns:
            Deduplicated DataFrame
        """
        # Simple deduplication based on failure mode similarity
        # In production, could use more sophisticated NLP similarity
        
        logger.info("Removing duplicate failure modes...")
        
        # Group by similar failure modes (simple text matching)
        # Columns are already Title Case here — _format_output is called per-source
        # inside generate_from_text/generate_from_structured before reaching this point
        fmea_df['failure_mode_lower'] = fmea_df['Failure Mode'].str.lower().str.strip()
        
        # Keep the entry with highest RPN for each similar failure
        deduplicated = fmea_df.sort_values('Rpn', ascending=False).drop_duplicates(
            subset=['failure_mode_lower'], keep='first'
        )
        
        deduplicated = deduplicated.drop(columns=['failure_mode_lower'])
        
        removed_count = len(fmea_df) - len(deduplicated)
        if removed_count > 0:
            logger.info(f"Removed {removed_count} duplicate entries")
        
        return deduplicated
    
    def export_fmea(self, fmea_df: pd.DataFrame, output_path: str, 
                   format: str = 'excel'):
        """
        Export FMEA to file
        
        Args:
            fmea_df: FMEA DataFrame to export
            output_path: Output file path
            format: 'excel' or 'csv'
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        if format.lower() == 'excel':
            # Export to Excel with formatting
            with pd.ExcelWriter(output_path, engine='xlsxwriter') as writer:
                fmea_df.to_excel(writer, sheet_name='FMEA', index=False)
                
                # Get workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['FMEA']
                
                # Add formats
                header_format = workbook.add_format({
                    'bold': True,
                    'bg_color': '#4472C4',
                    'font_color': 'white',
                    'border': 1
                })
                
                # Format headers
                for col_num, value in enumerate(fmea_df.columns.values):
                    worksheet.write(0, col_num, value, header_format)
                
                # Auto-adjust column widths
                for i, col in enumerate(fmea_df.columns):
                    max_length = max(
                        fmea_df[col].astype(str).apply(len).max(),
                        len(col)
                    )
                    worksheet.set_column(i, i, min(max_length + 2, 50))
            
            logger.info(f"FMEA exported to Excel: {output_path}")
            
        else:  # CSV
            fmea_df.to_csv(output_path, index=False)
            logger.info(f"FMEA exported to CSV: {output_path}")

    # ------------------------------------------------------------------ #
    #  Excel export with automated disagreement highlighting              #
    # ------------------------------------------------------------------ #

    def export_fmea_with_alerts(self, fmea_df: pd.DataFrame,
                                output_path: str,
                                consensus_threshold: float = 0.5) -> str:
        """
        Export FMEA to Excel with conditional formatting:
        - Rows where Consensus_Status == '🔴 HIGH DISAGREEMENT' get a red fill.
        - Rows where Review_Required == 'YES' get a bold red font.
        - The header row gets a branded blue style.
        - Columns are auto-sized.

        The DataFrame is expected to contain the columns
        ``Disagreement_Score``, ``Consensus_Status`` and ``Review_Required``
        (call ``enrich_with_consensus`` in analytics.py first).

        Args:
            fmea_df: Enriched FMEA DataFrame.
            output_path: Destination .xlsx path.
            consensus_threshold: Unused here (threshold is baked into the
                Consensus_Status column upstream).  Kept for API symmetry.

        Returns:
            The resolved output path string.
        """
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        output_path = str(Path(output_path))
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)

        # --- Step 1: plain export ------------------------------------------------
        fmea_df.to_excel(output_path, index=False, engine='openpyxl')

        # --- Step 2: open & style ------------------------------------------------
        wb = load_workbook(output_path)
        ws = wb.active

        # Style definitions
        red_fill   = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        amber_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        red_font   = Font(color='9C0006', bold=True)
        amber_font = Font(color='9C6500')
        green_font = Font(color='006100')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        # --- Header row ----------------------------------------------------------
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border

        # --- Locate the status column (if present) --------------------------------
        col_names = [c.value for c in ws[1]]
        status_idx = None
        review_idx = None
        score_idx = None
        if 'Consensus_Status' in col_names:
            status_idx = col_names.index('Consensus_Status') + 1  # 1-based
        if 'Review_Required' in col_names:
            review_idx = col_names.index('Review_Required') + 1
        if 'Disagreement_Score' in col_names:
            score_idx = col_names.index('Disagreement_Score') + 1

        # --- Row-level highlighting -----------------------------------------------
        for row_num in range(2, ws.max_row + 1):
            status_val = ws.cell(row=row_num, column=status_idx).value if status_idx else None
            review_val = ws.cell(row=row_num, column=review_idx).value if review_idx else None

            if status_val and 'HIGH DISAGREEMENT' in str(status_val):
                fill, font = red_fill, red_font
            elif status_val and 'MEDIUM' in str(status_val):
                fill, font = amber_fill, amber_font
            else:
                fill, font = green_fill, green_font

            for cell in ws[row_num]:
                cell.fill = fill
                cell.font = font
                cell.border = thin_border

            # Extra emphasis on the Review_Required cell
            if review_val == 'YES' and review_idx:
                cell_rev = ws.cell(row=row_num, column=review_idx)
                cell_rev.font = Font(color='9C0006', bold=True, size=12)

        # --- Auto-size columns ----------------------------------------------------
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_len = max(max_len, len(str(cell.value or '')))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 3, 55)

        # --- Freeze header row ----------------------------------------------------
        ws.freeze_panes = 'A2'

        wb.save(output_path)
        logger.info(f"Exported styled FMEA with alerts to {output_path}")
        return output_path


if __name__ == "__main__":
    # Example usage
    import yaml
    
    with open('../config/config.yaml', 'r') as f:
        config = yaml.safe_load(f)
    
    generator = FMEAGenerator(config)
    
    # Test with sample text
    sample_texts = [
        "The engine failed completely after 50k miles. This caused the car to stop on the highway, creating a dangerous situation.",
        "Brake system malfunction - brakes became unresponsive during heavy rain. Almost caused an accident."
    ]
    
    fmea = generator.generate_from_text(sample_texts, is_file=False)
    print("\nGenerated FMEA:")
    print(fmea)
